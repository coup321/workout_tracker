'''

This is a helper script to support google drive integration into the workout_tracking app.

The app functions with auth/drive/file permissions so it does not require verification
the consequence of this is that we simply must upload our file from this app to the drive
in order to have access to it.

There are three primary functions:
1)  get_file(id): downloads and returns the file 'id'

2)  add_file('name'): adds an excel file with name = 'name' to the working directory and
    to google drive. It prints the file ID which should be saved and addended to the
    file_id variable in this file.
    WARNING: this overwrites existing files without warning

3)  build_service: handles credentials and builds the google drive API driver service.
    This function requires a file called credentials.json to be in the working 
    directory. This JSON has necessary authentication information. The function
    will open a web browser and facilitate authentication and creation of a
    tolken file (which also must remain in the working directory) which
    allows concurrent access to the drive with an abbreviated auth process.





'''

from __future__ import print_function
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient.http import MediaFileUpload
import io
from openpyxl import Workbook

#represents google drive file ID
#if None should be manually updated or run add_file('name')
my_file_id = '19jWuQSbnjX4wE78SjCW-Cfy1tWtx874g'

# If modifying these scopes, delete the file token.pickle.
SCOPES = [
        'https://www.googleapis.com/auth/drive.metadata.readonly', 
        'https://www.googleapis.com/auth/drive.file']

def build_service():
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server()
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)
    return build('drive', 'v3', credentials=creds)

def get_file(file_id=my_file_id):
    #assures that build_service has run, and then retrives file from drive
    try:
        creds
    except NameError:
        service = build_service()

    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
    return fh

def add_file(name):
    #adds xlsx file to working drive, google drive, updates file_id, and prints file_id
    #WARNING: this operation will overwrite existing
    #files without warning!!!
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet 1'
    ws.cell(row=1,column=1,value = 'Date')
    ws.cell(row=1,column=2,value = 'Exercise')
    ws.cell(row=1,column=3,value = 'Sets')
    wb.save(name)

    try:
        creds
    except NameError:
        service = build_service()

    file_metadata = {'name': name}
    mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    media = MediaFileUpload(name,
                            mimetype=mime)
    file = service.files().create(body=file_metadata,
                                        media_body=media,
                                        fields='id').execute()
    print('File ID: %s' % file.get('id'))
    return 0
add_file('test1.xlsx')
